#!/usr/bin/env python3
"""
Exhibit 509 — scorched-earth rebuild, step 1: inventory + schema probe.

Walks sources/509/<year>/, and for every workbook it:
  - records sha256, size, year (from the folder), and a guessed ABA section
    (from the filename) into pipeline/manifest.csv  [committed, byte-pins inputs]
  - probes the structure (sheet names; for each sheet the dims + first header-ish
    rows) into pipeline/schema_probe/<year>__<file>.txt  [so we can see the real
    layout before writing a single extractor — the ABA format drifts by year]

Handles modern .xlsx (openpyxl) and legacy .xls (xlrd 2.0). Read-only; never
touches the source files. Run: python3 pipeline/inspect_sources.py
"""
import csv
import hashlib
import os
import re
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "sources", "509")
MANIFEST = os.path.join(ROOT, "pipeline", "manifest.csv")
PROBE_DIR = os.path.join(ROOT, "pipeline", "schema_probe")

# Filename keyword -> canonical ABA 509 section. The disclosures ship one
# workbook per section; names vary year to year, so match loosely + log misses.
SECTION_HINTS = [
    ("admission", "admissions"), ("lsat", "admissions"), ("gpa", "admissions"),
    ("enroll", "enrollment"), ("jd", "enrollment"),
    ("attrition", "attrition"),
    ("curricul", "curriculum"), ("clinic", "curriculum"),
    ("faculty", "faculty"), ("librar", "faculty"),
    ("employ", "employment"), ("outcome", "employment"),
    ("bar", "bar_passage"), ("pass", "bar_passage"),
    ("tuition", "tuition"), ("fee", "tuition"), ("living", "tuition"),
    ("grant", "grants"), ("scholar", "grants"), ("conditional", "grants"),
    ("transfer", "transfers"),
]


def guess_section(name):
    low = name.lower()
    for kw, sec in SECTION_HINTS:
        if kw in low:
            return sec
    return "UNKNOWN"


def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def probe_xlsx(path, out):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        try:
            dims = ws.calculate_dimension()
        except Exception:
            dims = "?"
        out.write(f"\n## sheet: {ws.title!r}  dims={dims} "
                  f"max_row={ws.max_row} max_col={ws.max_column}\n")
        for i, row in enumerate(ws.iter_rows(max_row=8, values_only=True)):
            cells = ["" if c is None else str(c) for c in row]
            out.write(f"  r{i}: {cells}\n")
    wb.close()


def probe_xls(path, out):
    import xlrd
    wb = xlrd.open_workbook(path)
    for ws in wb.sheets():
        out.write(f"\n## sheet: {ws.name!r}  nrows={ws.nrows} ncols={ws.ncols}\n")
        for i in range(min(8, ws.nrows)):
            cells = [str(ws.cell_value(i, j)) for j in range(ws.ncols)]
            out.write(f"  r{i}: {cells}\n")


def main():
    if not os.path.isdir(SRC):
        sys.exit(f"no source tree at {SRC}")
    os.makedirs(PROBE_DIR, exist_ok=True)
    rows = []
    for year in sorted(os.listdir(SRC)):
        ydir = os.path.join(SRC, year)
        if not os.path.isdir(ydir):
            continue
        for fn in sorted(os.listdir(ydir)):
            ext = os.path.splitext(fn)[1].lower()
            if ext not in (".xlsx", ".xls", ".xlsb"):
                continue
            path = os.path.join(ydir, fn)
            section = guess_section(fn)
            rows.append({
                "year": year, "section": section, "file": fn,
                "ext": ext, "bytes": os.path.getsize(path), "sha256": sha256(path),
            })
            probe_path = os.path.join(PROBE_DIR, f"{year}__{re.sub(r'[^A-Za-z0-9._-]', '_', fn)}.txt")
            with open(probe_path, "w", encoding="utf-8") as out:
                out.write(f"# {fn}  (year={year}, guessed section={section})\n")
                try:
                    (probe_xlsx if ext in (".xlsx", ".xlsb") else probe_xls)(path, out)
                except Exception as e:
                    out.write(f"\n!! probe failed: {type(e).__name__}: {e}\n")

    rows.sort(key=lambda r: (r["year"], r["section"], r["file"]))
    with open(MANIFEST, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["year", "section", "file", "ext", "bytes", "sha256"])
        w.writeheader()
        w.writerows(rows)

    # console summary: coverage grid (year x section) + unknowns
    years = sorted({r["year"] for r in rows})
    secs = sorted({r["section"] for r in rows})
    print(f"{len(rows)} workbooks across {len(years)} years -> {MANIFEST}\n")
    print("coverage (year x section, count):")
    hdr = "year  " + "  ".join(s[:8].rjust(8) for s in secs)
    print(hdr)
    for y in years:
        cells = []
        for s in secs:
            c = sum(1 for r in rows if r["year"] == y and r["section"] == s)
            cells.append((str(c) if c else "·").rjust(8))
        print(f"{y}  " + "  ".join(cells))
    unk = [r for r in rows if r["section"] == "UNKNOWN"]
    if unk:
        print(f"\n{len(unk)} files with UNKNOWN section (need a SECTION_HINTS rule):")
        for r in unk:
            print(f"  {r['year']}/{r['file']}")
    print(f"\nschema probes -> {PROBE_DIR}/")


if __name__ == "__main__":
    main()
