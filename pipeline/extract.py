#!/usr/bin/env python3
"""
Exhibit 509 rebuild, step 3: extract workbooks -> tidy facts (SQLite).

Each ABA section is a tidy one-row-per-school sheet. We register, per section,
the source column header for every gz field we can read *directly* (no math),
plus a cleaner. Derived fields (percentages, firm-size buckets, schol_none) are
intentionally NOT here — they belong to the rebuild/derivation step; keeping the
extract layer to raw reads is what makes the oracle diff meaningful.

Output: pipeline/facts.sqlite, table `facts`
  (school_id, year, section, field, value, src_file, sheet, row, col)

Usage: python3 pipeline/extract.py [year ...]   (default: all years present)
"""
import glob
import os
import re
import sqlite3
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from crosswalk import build_resolver

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "sources", "509")
DB = os.path.join(ROOT, "pipeline", "facts.sqlite")

SENT_STR = {"", "n/a", "na", "n", "*", "**", "***", "-", "--"}


def clean_int(v):
    if v is None:
        return None
    s = str(v).strip().replace(",", "").replace("$", "")
    if s.lower() in SENT_STR:
        return None
    try:
        return int(round(float(s)))
    except ValueError:
        return None


def clean_money(v):  # tuition/grant $ strings; 0 is a real-but-suspect value, keep
    return clean_int(v)


def clean_float(v):
    if v is None:
        return None
    s = str(v).strip().replace(",", "").replace("%", "").replace("$", "")
    if s.lower() in SENT_STR:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def clean_pct(v):
    return clean_float(v)


def clean_zeronull(v):  # LSAT/uGPA/grant-GPA: stored 0 means "not reported"
    f = clean_float(v)
    return None if f == 0 else f


def clean_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s.lower() in SENT_STR else s


def clean_yn(v):
    s = clean_str(v)
    if s is None:
        return None
    return s[0].upper() if s[:1].upper() in ("Y", "N") else s


# ── section registry ──────────────────────────────────────────────────────────
# file_glob matched within sources/509/<year>/. name_col defaults to col 0.
# fields: gz_field -> (source header, cleaner). Headers matched case/space-loose.
SECTIONS = {
    "admissions": {
        "glob": "First_Year_Class*",
        "fields": {
            "apps": (["Applications", "CompletedApplications", "Completed Apps",
                      "# Of Apps Total"], clean_int),
            "offers": (["Offers", "OffersAdmission", "Offer", "# Of Offers Total"], clean_int),
            "acc": (["AcceptanceRate", "Acceptence Rate", "Acceptance Rate"], clean_pct),
            "enr_1l": (["TotalEnrollees", "TotalFYClassAll", "Total FY class ALL",
                        "Total FY class", "# Of Matriculants Total"], clean_int),
            "enr_1l_entering": (["Enrollees", "EnrolleesFromApplicantPool",
                                 "Enrollees from App pool"], clean_int),
            "enr_1l_ft": (["FTEnrollees", "TotalFYClassFT", "Total FY class FT",
                           "# Of Matriculants Full-Time"], clean_int),
            "enr_1l_pt": (["PTEnrollees", "TotalFYClassPT", "Total FY class PT",
                           "# Of Matriculants Part-Time"], clean_int),
            "gpa75": (["All75thPercentileUGPA", "All75GPA", "75th percentile UGPA ALL",
                       "75th percentile UGPA", "75th Percentile GPA Total"], clean_zeronull),
            "gpa50": (["All50thPercentileUGPA", "All50GPA", "50th percentile UGPA ALL",
                       "50th percentile UGPA", "50th Percentile GPA Total"], clean_zeronull),
            "gpa25": (["All25thPercentileUGPA", "All25GPA", "25th percentile UGPA ALL",
                       "25th percentile UGPA", "25th Percentile GPA Total"], clean_zeronull),
            "lsat75": (["All75thPercentileLSAT", "All75LSAT", "75th percentile LSAT ALL",
                        "75th percentile LSAT", "75th Percentile LSAT Total"], clean_zeronull),
            "lsat50": (["All50thPercentileLSAT", "All50LSAT", "50th percentile LSAT ALL",
                        "50th percentile LSAT", "50th Percentile LSAT Total"], clean_zeronull),
            "lsat25": (["All25thPercentileLSAT", "All25LSAT", "25th percentile LSAT ALL",
                        "25th percentile LSAT", "25th Percentile LSAT Total"], clean_zeronull),
            "gre_takers": ("GRETotalEnrollees", clean_int),  # absent pre-2023
        },
    },
    "faculty": {
        "glob": "Faculty_Resources*",
        "fields": {
            "fac_ft": (["FTTotal", "Total FT", "Full-Time Fall Total"], clean_int),
            "fac_pt": (["NONFTTotal", "Total NonFT", "Part-Time Fall Total"], clean_int),
            "fac_total": (["TotalFaculties", "Total", "Total Fall"], clean_int),
            "fac_men": (["MaleTotal", "Total Male", "Total Fall Men"], clean_int),
            "fac_women": (["FemaleTotal", "Total Female", "Total Fall Women"], clean_int),
            "fac_poc": (["POCTotal", "Total People of Color", "Total Minority",
                         "Total Fall Minorities"], clean_int),
            "librarians_total": ("TotalLibrarians", clean_int),
        },
    },
    "transfers": {
        "glob": "Transfers*",
        "fields": {
            "trans_out": (["JD1 Transfers Out", "1L Transfers Out", "Transfers Out"], clean_int),
            "trans_in": (["TransferIn", "Transfer In", "Transfers In",
                          "Total # of Transfer in Students"], clean_int),
            "trans_gpa75": (["75th Percentile JD1 GPA",
                             "75th percentile 1L GPA (12 or more transfers in)",
                             "75th Percentile 1L GPA"], clean_zeronull),
            "trans_gpa50": (["50th Percentile JD1 GPA", "GPA50thPercentile",
                             "50th percentile 1L GPA (12 or more transfers in)",
                             "50th Percentile 1L GPA"], clean_zeronull),
            "trans_gpa25": (["25th Percentile JD1 GPA", "GPA25thPercentile",
                             "25th percentile 1L GPA (12 or more transfers in)",
                             "25th Percentile 1L GPA"], clean_zeronull),
        },
    },
    "bar_first": {
        "glob": "First_Time_Bar*",
        "name_col_alt": "School Name",
        "fields": {
            "bar_grads": (["Graduates In {Y-1}", "Total Graduates"], clean_int),
            "bar_first_takers": ("Total First Time Takers", clean_int),
            "bar_first_passers": ("Total First Time Passers", clean_int),
            "bar": ("AvgSchoolPassPercent*", clean_pct),
            "bar_state_avg": ("AvgStatePassPercent**", clean_pct),
            "bar_state_diff": ("TotalDifferencePercent***", clean_pct),
        },
    },
    # ≤2016 single bar-passage sheet (replaced by First/Two-Year split in 2018+).
    # Multiple rows per school (exam-year × jurisdiction) + embedded prior report
    # years, so filter to this report year; school-level Composite columns are
    # identical across a school's jurisdiction rows (dedup via collision skip).
    "bar_rates": {
        "glob": "Bar_Passage_Rates*",
        "name_col_alt": "School Name",
        "row_filter": ("Reporting Year", "{Y}"),
        "fields": {
            "bar": ("Composite Avg. School Pass %", clean_pct),
            "bar_state_avg": ("Composite Avg. State Pass %", clean_pct),
            "bar_state_diff": ("Composite Avg. Pass Diff. %", clean_pct),
            "bar_first_takers": ("Total First-Time Takers", clean_int),
        },
    },
    "conditional": {
        "glob": "Conditional_Scholarships*",
        "fields": {
            "cond_enter": ("{Y-1}-{Y} # Entering With", clean_int),
            "cond_elim": ("{Y-1}-{Y} # Eliminated", clean_int),
        },
    },
    "bar_2yr": {
        "glob": "TwoYear_Ultimate_Bar*",
        "name_col_alt": "School Name",
        "fields": {
            "bar_2yr_grads": (["{Y-3} Graduates", "No. of Graduates"], clean_int),
            "bar_2yr_takers": (["{Y-3} Takers", "No. of Takers"], clean_int),
            "bar_2yr_passers": (["{Y-3} Passers", "No. of Passers"], clean_int),
            "bar_2yr": ("%Passers", clean_pct),
        },
    },
    "grants": {
        "glob": "Grants_and_Scholarships*",
        "fields": {
            "schol_total": (["Total Number # of Recieving Grants Total #",
                             "Total # receiving grants total #",
                             "Total # receiving grants"], clean_int),
            "schol_lt": (["Less than half tuition Total Number #",
                          "Less than half tuition total #",
                          "Total Less than 1/2 tuition"], clean_int),
            "schol_mt": (["Half to full tuition total Number #",
                          "Half to full tuition total #",
                          "Total Half to full tuition"], clean_int),
            "schol_full": (["Full tuition total Number #",
                            "Full tuition total #", "Total Full tuition"], clean_int),
            "schol_gt": (["More than full tuition total Number #",
                          "More than full tuition total #",
                          "Total More than full tuition"], clean_int),
            "grant_med_ft": (["FT 50th percentile grant amount",
                              "Full-Time 50th Per centile grant amount"], clean_money),
            "grant_p25_ft": (["FT 25th percentile grant amount",
                              "Full-Time 25th Percentile grant amount"], clean_money),
            "grant_p75_ft": (["FT 75th percentile grant amount",
                              "Full-Time 75th Percentile grant amount"], clean_money),
        },
    },
    "employment": {
        "glob": "Employment_Summary*",
        "fields": {
            # 2023 used the 'Number' suffix + 'BarPassageRequired'; 2024+ use
            # 'Total' + 'BarAdmissionRequired'.
            "emp_grads": (["Total_GraduatesTotal", "Total_GraduatesNumber"], clean_int),
            "ftlt": ("Total_FTLT", clean_int),
            "emp_bar_ftlt": (["Employed_BarAdmissionRequiredFTLT",
                              "Employed_BarPassageRequiredFTLT"], clean_int),
            "emp_jda_ftlt": ("Employed_JDAdvantageFTLT", clean_int),
            "emp_solo_ftlt": ("Solo-FTLT", clean_int),
            "emp_seeking": (["UnEmployedSeekingTotal", "UnEmployedSeekingNumber"], clean_int),
            "emp_not_seeking": (["UnEmployedNotSeekingTotal",
                                 "UnEmployedNotSeekingNumber"], clean_int),
        },
    },
    "curriculum": {
        "glob": "Curricular_Offerings*",
        "fields": {
            # pre-2025 reported a single count (= filled); 2025 split into
            # Available + Filled, and dropped the Seminars column.
            "clinics_available": (["LawClinicsAvailable",
                                   "# of clinic seats available"], clean_int),
            "clinics_filled": (["LawClinicsFilled", "LawClinics", "ClincicSum"], clean_int),
            "field_placements_filled": (["FieldPlacementsFilled", "FieldPlacements",
                                         "FieldPlacementSum",
                                         "# of field placement positions filled"], clean_int),
            "sim_courses_available": (["SimulationCoursesAvailable",
                                       "# of simulation course seats available"], clean_int),
            "sim_courses_filled": (["SimulationCoursesFilled", "SimulationCourses",
                                    "SimulationSum"], clean_int),
            "seminars": (["Seminars", "Seminarcount"], clean_int),
            "co_curricular": (["CoCurricularOfferings", "CocurricularCount"], clean_int),
        },
    },
    "attrition": {
        "glob": "Attrition*",
        "fields": {
            "atr_acad_1l": (["AcademicAttrition_TotalJD1Total",
                             "AcadAttrition_TotalJD1Total"], clean_int),
            "atr_acad_1l_pct": (["AcademicAttrition_TotalJD1Percentage",
                                 "AcadAttrition_TotalJD1Percentage"], clean_pct),
            "atr_acad_ul_pct": (["AcademicAttrition_TotalULPercentage",
                                 "AcadAttrition_TotalULPercentage"], clean_pct),
            "atr_other_1l": ("OtherAttrition_TotalJD1Total", clean_int),
            "atr_other_1l_pct": ("OtherAttrition_TotalJD1Percentage", clean_pct),
            "atr_other_ul_pct": ("OtherAttrition_TotalULPercentage", clean_pct),
        },
    },
    "enrollment": {
        "glob": "JD_Enrollment_and_Ethnicity*",
        "fields": {
            # 2017+ report combined "Grand Total" columns; 2011-2016 report a
            # "<x> Total" column (= Men+Women summed) alongside the #/%-per-sex
            # breakdown. Both era's headers are listed so one extractor covers all.
            "enr": (["TotalGrandTotal", "Total Grand Total", "#Total Total"], clean_int),
            "grads": (["Total Degrees Awarded", "#Total J.D. Deg Awd"], clean_int),
            "race_white": (["WhiteGrandTotal", "White Grand Total", "#White Total"], clean_int),
            "race_black": (["BlackGrandTotal", "Black or African American Grand Total",
                            "#Black or African American Total"], clean_int),
            "race_hisp": (["HispGrandTotal", "OtherHispGrandTotal", "Hispanic Grand Total",
                           "#Hispanics of any race Total"], clean_int),
            "race_asian": (["AsianGrandTotal", "Asian Grand Total", "#Asian Total"], clean_int),
            "race_indian": (["AmericanIndianGrandTotal", "AmerIndian Grand Total",
                             "#American Indian or Alaska Native Total"], clean_int),
            "race_native": (["NativeGrandTotal",
                             "Native Hawaiian Pacific Islander Grand Total",
                             "#Native Hawaiian or Other Pacific Islander Total"], clean_int),
            "race_multi": (["MultiracialGrandTotal", "RaceGrandTotal",
                            "Two or more Races Grand Total", "#Two or more races Total"], clean_int),
            "race_nr": (["NRGrandTotal", "NonRes Alien Grand Total",
                         "#Nonresident Alien Total"], clean_int),
            "race_unknown": (["UnknownGrandTotal", "UnknownRaceGrandTotal",
                              "Race Unk Grand Total", "#Race and Ethnicity Unknown Total"], clean_int),
            # sex grand totals: 2011-2016 expose '#Total Men'/'#Total Women' directly;
            # 2017+ have no sex subtotal column (gz already carries 2018-2026), so these
            # backfill 2011-2016 only.
            "sex_men": ("#Total Men", clean_int),
            "sex_women": ("#Total Women", clean_int),
        },
    },
    "basics": {
        "glob": "The_Basics*",
        "fields": {
            "school_type": (["SchoolType", "Type of School"], clean_str),
            "app_fee": (["AppFee", "Application Fee"], clean_money),
            "term": ("Term", clean_str),
            "credit_hours_required": (["RequiredCreditHours", "# of Credit Hours for JD"], clean_int),
        },
    },
    "tuition": {
        "glob": "Tuitions_and_Fees*",
        "fields": {
            "tui_ft_res": (["FT_Resident_Annual", "FT_Resident_Semester",
                            "Full Time Resident Semester", "Full Time Resident",
                            "Full-Time Resident"], clean_money),
            "tui_ft_nonres": (["FT_NonResident_Annual", "FT_NonResident_Semester",
                               "Full Time Non resident Semester", "Full Time Non resident",
                               "Full-Time Non-Resident"], clean_money),
            "tui_pt_res": (["PT_Resident_Annual", "PT_Resident_Semester",
                            "Part Time Resident Semester", "Part Time Resident",
                            "Part-Time Resident"], clean_money),
            "tui_pt_nonres": (["PT_NonResident_Annual", "PT_NonResident_Semester",
                               "Part Time Non resident Semester", "Part Time Non resident",
                               "Part-Time Non-Resident"], clean_money),
            "ft_fee": (["FTRS_AnnualFees", "FTRS Annual Fees", "FTRS Fees"], clean_money),
            "living_on_campus": (["Living_On_Campus", "Living On Campus", "Living on Campus"], clean_money),
            "living_off_campus": (["Living_Off_Campus", "Living Off Campus"], clean_money),
            "living_at_home": (["Living_At_Home", "Living At Home", "Living at Home"], clean_money),
            # stored raw to mirror the gz, which is faithful to each year's source
            # ('Yes'/'No' in 2023, 'Y'/'N' in 2024+); column absent pre-2023.
            "cond_offered": ("OfferScholorships", clean_str),
        },
    },
}


COLLISIONS = []  # (year, section, sid, field, kept_name, kept_val, dropped_name, dropped_val)

# The canonical tui_* fields are ANNUAL. Through 2020 the ABA reported tuition
# PER SEMESTER (verified: gz/source ratio is exactly 2.0 for every non-zero
# 2020 value); 2021+ report annual. So for these years we annualize (×2) the
# four tuition fields on extract. Earlier years get added here as the oracle /
# curated trends confirm them.
PER_SEMESTER_TUITION_YEARS = {2018, 2019, 2020}
ANNUALIZE_FIELDS = {"tui_ft_res", "tui_ft_nonres", "tui_pt_res", "tui_pt_nonres"}

# fields that legitimately exist only in some years' workbooks — a missing header
# here is expected, not a problem, so it is not warned about.
OPTIONAL_FIELDS = {"race_nr", "enr_1l_entering",
                   "clinics_available", "sim_courses_available", "seminars",
                   "cond_offered",  # cond_offered column absent pre-2023
                   "gre_takers",    # GRE takers count absent pre-2023
                   "ft_fee",        # fee column absent some years (e.g. 2018)
                   # absent in the older (≤2017) workbook structures:
                   "enr_1l_ft", "enr_1l_pt", "librarians_total",
                   "clinics_filled", "sim_courses_filled", "co_curricular",
                   "field_placements_filled",  # FT/PT-split (no combined) in ≤2014
                   "atr_acad_1l", "atr_acad_1l_pct", "atr_acad_ul_pct",
                   "atr_other_1l", "atr_other_1l_pct", "atr_other_ul_pct",
                   "trans_out", "credit_hours_required", "acc",
                   "trans_gpa75", "trans_gpa50", "trans_gpa25",  # absent ≤2013
                   # ≤2016 enrollment is #/%-per-sex with no per-race grand totals;
                   # race/enr totals are deferred (need aggregation, not trend-checked):
                   "enr", "race_white", "race_black", "race_hisp", "race_asian",
                   "race_indian", "race_native", "race_multi", "race_nr", "race_unknown",
                   "sex_men", "sex_women",  # '#Total Men/Women' only in 2011-2016 sheets
                   "grant_med_ft", "grant_p25_ft", "grant_p75_ft"}  # renamed ≤2011


def hkey(s):
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


def candidates(src, year):
    """A field's source header may be a str or a list of fallbacks (ABA renames
    columns across years). Tokens {Y-1}/{Y-3} expand to cohort years (e.g. the
    two-year-ultimate bar cohort is report-year minus 3)."""
    opts = src if isinstance(src, (list, tuple)) else [src]
    out = []
    for o in opts:
        o = (o.replace("{Y-1}", str(year - 1)).replace("{Y-3}", str(year - 3))
              .replace("{Y}", str(year)))
        out.append(o)
    return out


def extract_year(year, resolve, conn):
    ydir = os.path.join(SRC, str(year))
    if not os.path.isdir(ydir):
        return 0
    rows_out = 0
    for section, spec in SECTIONS.items():
        matches = glob.glob(os.path.join(ydir, spec["glob"] + ".xls*"))
        if not matches:
            continue
        path = matches[0]
        fname = os.path.basename(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        header = [hkey(c) for c in rows[0]]
        hidx = {h: i for i, h in enumerate(header)}
        # optional row filter (header, value) — e.g. keep only this report year
        rf = spec.get("row_filter")
        rf_idx, rf_val = None, None
        if rf:
            rf_idx = hidx.get(hkey(rf[0]))
            rf_val = hkey(candidates(rf[1], int(year))[0])
        # resolve each field's column index once
        colmap = {}
        for gzf, (src, cleaner) in spec["fields"].items():
            i = next((hidx[hkey(c)] for c in candidates(src, int(year))
                      if hkey(c) in hidx), None)
            if i is None:
                if gzf not in OPTIONAL_FIELDS:
                    print(f"  !! {section}: header {src!r} not found in {fname}")
                continue
            colmap[gzf] = (i, cleaner)
        seen = {}  # (sid, field) -> (value, source_name) within this sheet
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            sid = resolve(r[0])
            if not sid:
                continue  # footnote / unmatched (reported by crosswalk probe)
            if rf_idx is not None and (rf_idx >= len(r) or hkey(r[rf_idx]) != rf_val):
                continue  # row filtered out (wrong report year)
            for gzf, (i, cleaner) in colmap.items():
                if i >= len(r):
                    continue
                val = cleaner(r[i])
                if val is None:
                    continue
                if (int(year) in PER_SEMESTER_TUITION_YEARS
                        and gzf in ANNUALIZE_FIELDS and isinstance(val, (int, float))):
                    val = val * 2  # per-semester -> annual
                key = (sid, gzf)
                if key in seen and str(seen[key][0]) != str(val):
                    # two distinct source rows collapse to one slug — FLAG, keep
                    # first (canonical), record the collision for adjudication.
                    COLLISIONS.append(
                        (int(year), section, sid, gzf,
                         seen[key][1], seen[key][0], r[0], val))
                    continue
                if key in seen:
                    continue
                seen[key] = (val, r[0])
                conn.execute(
                    "INSERT INTO facts VALUES (?,?,?,?,?,?,?,?,?)",
                    (sid, int(year), section, gzf,
                     str(val), fname, ws.title, None, i),
                )
                rows_out += 1
        wb.close()
    return rows_out


def extract_2017_sex(resolve, conn):
    """2017 is the one JD-enrollment year with neither a '#Total Men' column
    (2011-2016) nor gz-carried sex (2018-2026). Its sheet does expose per-JD-level
    'Total {n}L Men/Women' subtotals, so sex_{men,women} = sum of the three levels."""
    matches = glob.glob(os.path.join(SRC, "2017", "JD_Enrollment_and_Ethnicity*.xls*"))
    if not matches:
        return 0
    men_cols = [hkey(c) for c in ("Total 1L Men", "Total 2L Men", "Total 3L Men")]
    wom_cols = [hkey(c) for c in ("Total 1L Women", "Total 2L Women", "Total 3L Women")]
    wb = openpyxl.load_workbook(matches[0], read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hidx = {hkey(c): i for i, c in enumerate(rows[0])}
    fname = os.path.basename(matches[0])
    n = 0
    for field, cols in (("sex_men", men_cols), ("sex_women", wom_cols)):
        idxs = [hidx[c] for c in cols if c in hidx]
        if len(idxs) != 3:
            print(f"  !! 2017 sex: expected 3 level cols for {field}, found {len(idxs)}")
            continue
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            sid = resolve(r[0])
            if not sid:
                continue
            vals = [clean_int(r[i]) for i in idxs if i < len(r)]
            vals = [v for v in vals if v is not None]
            if not vals:
                continue
            conn.execute("INSERT INTO facts VALUES (?,?,?,?,?,?,?,?,?)",
                         (sid, 2017, "enrollment", field, str(sum(vals)),
                          fname, ws.title, None, None))
            n += 1
    wb.close()
    print(f"2017 sex aggregation: {n} facts")
    return n


def main():
    years = sys.argv[1:] or sorted(
        d for d in os.listdir(SRC) if d.isdigit() and os.path.isdir(os.path.join(SRC, d))
    )
    resolve, _ = build_resolver()
    conn = sqlite3.connect(DB)
    conn.execute("DROP TABLE IF EXISTS facts")
    conn.execute(
        "CREATE TABLE facts (school_id TEXT, year INT, section TEXT, field TEXT, "
        "value TEXT, src_file TEXT, sheet TEXT, row INT, col INT)"
    )
    total = 0
    for y in years:
        n = extract_year(y, resolve, conn)
        print(f"{y}: {n} facts")
        total += n
    if not sys.argv[1:] or "2017" in years:
        total += extract_2017_sex(resolve, conn)
    conn.commit()
    print(f"\n{total} facts -> {DB}")
    if COLLISIONS:
        seen_pairs = {(c[2], c[6]) for c in COLLISIONS}
        print(f"\n!! {len(COLLISIONS)} id collisions (two source rows -> one slug) "
              f"across {len(seen_pairs)} school-pairs — FLAG for adjudication:")
        for yr, sec, sid, fld, kn, kv, dn, dv in COLLISIONS[:12]:
            print(f"   {yr} {sid}.{fld}: kept {kn!r}={kv} / dropped {dn!r}={dv}")
    conn.close()


if __name__ == "__main__":
    main()
